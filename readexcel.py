import re
from openpyxl import load_workbook


wb = load_workbook("20250320.xlsx")

sheet_mapping = {
    "BAY2-BS.4": "meter-4",
    "BAY3-PL.8": "meter-8",
    "BAY4-BS.7": "meter-7",
    "BAY4-BS.11": "meter-11",
    "BAY6-PL.16": "meter-16",
    "BAY7-BS.12": "meter-12",
    "BAY7-BS.15": "meter-15"
}

reference_data = {
    "meter-4": ["8934", "5678"],
    "meter-8": ["9190", "8434", "8438", "8013", "8774", "9787", "8335", "8437", "8099", "9164", "9477", "8545", "9190", "9273", "9201", "8544", "9957", "9336", "9284", "9674", "9367", "8544", "9957", "9336", "9284", "9674", "9367", "8544", "9316", "8446", "9507", "9325", "8770", "9171", "9367", "9774", "9812", "9766", "9731", "8545", "9347", "9787", "9445", "9673", "9787", "9719", "9749"],
    "meter-7": ["4444", "5555"],
    "meter-11": ["8548", "7777"],
    "meter-12": ["8136", "8548"],
    "meter-15": ["8097", "8040"],
    "meter-16": ["8614", "9201"]
}

start_row = 9 
truck_column = "C"

def extract_numbers(text):
    if text:
        numbers = re.findall(r"\d+", str(text)) 
        return "".join(numbers) if numbers else None 
    return None

truck_data = {}

# Iterasi melalui sheet yang dipilih
for sheet_name, meter_key in sheet_mapping.items():
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        # Ambil angka dari kolom C, mulai dari baris 9 ke bawah
        truck_numbers = [
            extract_numbers(sheet[f"{truck_column}{row}"].value)
            for row in range(start_row, sheet.max_row + 1)
            if sheet[f"{truck_column}{row}"].value is not None
        ]

        # Filter agar tidak ada nilai None
        truck_numbers = [num for num in truck_numbers if num]

        # Simpan hasil dengan key yang sesuai
        truck_data[meter_key] = truck_numbers

# Dictionary untuk menyimpan hasil perbandingan
comparison_result = {}

# Perbandingan data
for meter, trucks in reference_data.items():
    excel_trucks = set(truck_data.get(meter, []))  # Data dari Excel (set untuk efisiensi pencarian)
    reference_trucks = set(trucks)  # Data referensi

    # Mencari data yang ada di referensi tapi tidak ada di Excel
    missing_trucks = reference_trucks - excel_trucks
    # Mencari data yang ada di Excel tapi tidak ada di referensi
    additional_trucks = excel_trucks - reference_trucks

    comparison_result[meter] = {
        "missing": list(missing_trucks),
        "additional": list(additional_trucks)
    }

# Menemukan truk yang hilang di satu meter tetapi ada di meter lain
missing_found_elsewhere = []

for meter, result in comparison_result.items():
    for missing_truck in result["missing"]:
        # Cek apakah nomor truk yang hilang ditemukan di meter lain
        found_in_other_meter = [
            other_meter for other_meter, data in truck_data.items() if missing_truck in data and other_meter != meter
        ]
        
        if found_in_other_meter:
            missing_found_elsewhere.append(f"Nomor {missing_truck} mengisi di {meter} sedangkan antrian di {', '.join(found_in_other_meter)}")

# Cetak hasil perbandingan
print("Hasil Perbandingan:")
print(comparison_result)

# Cetak truk yang berpindah meter
print("\nTruk yang berpindah meter:")
for message in missing_found_elsewhere:
    print(message)
