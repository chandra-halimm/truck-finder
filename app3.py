from flask import Flask, render_template
import re
from openpyxl import load_workbook

app = Flask(__name__)

# Mapping sheet ke key yang diinginkan
sheet_mapping = {
    "BAY2-BS.4": "meter-4",
    "BAY3-PL.8": "meter-8",
    "BAY4-BS.7": "meter-7",
    "BAY4-BS.11": "meter-11",
    "BAY6-PL.16": "meter-16",
    "BAY7-BS.12": "meter-12",
    "BAY7-BS.15": "meter-15"
}

# Data referensi
reference_data = {
    "meter-4": ["9787", "9329", "8256", "9820", "8635", "8636", "8315", "8934", "8040"],

    "meter-8": [ "9373", "9957", "9347", "9329", "8097", "8099", "9324", "8136", "8470", "8335", "9331", "9653", "9488", "9473", "8040"],

    "meter-7": ["9787", "9273", "9338", "9812", "8545", "8436", "8438", "9201"],

    "meter-11": ["9787", "9273", "9338", "8255", "9812", "8545", "8436", "8438", "9473", "9201"],

    "meter-12": ["9674", "9341", "9643", "9730", "9316", "8616", "8440", "9345", "8546", "9336", "9477"],
    
    "meter-15": ["9674", "9341", "9643", "9327", "9730", "9750", "9324", "9316", "9316", "8450", "9488", "8440", "9345", "8546", "9336", "9477"],

    "meter-16": ["9787", "9731","9327", "9483", "9336", "9749", "9750", "9356", "8134", "8013", "9811", "8336", "9354", "9784", "9171", "8048"]
}

start_row = 9  
truck_column = "C"

def extract_numbers(text):
    """Ekstrak angka dari teks."""
    if text:
        numbers = re.findall(r"\d+", str(text))
        return "".join(numbers) if numbers else None
    return None

def process_excel():
    """Proses file Excel dan bandingkan dengan data referensi."""
    wb = load_workbook("20250322.xlsx")

    truck_data = {}
    
    for sheet_name, meter_key in sheet_mapping.items():
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            truck_numbers = [
                extract_numbers(sheet[f"{truck_column}{row}"].value)
                for row in range(start_row, sheet.max_row + 1)
                if sheet[f"{truck_column}{row}"].value is not None
            ]

            truck_numbers = [num for num in truck_numbers if num]

            truck_data[meter_key] = truck_numbers

    # Memisahkan solar dan pertalite
    solar_meters = {key: value for key, value in truck_data.items() if key in ["meter-4", "meter-12", "meter-15"]}
    pertalite_meters = {key: value for key, value in truck_data.items() if key in ["meter-8", "meter-16"]}

    solar_reference = {key: value for key, value in reference_data.items() if key in solar_meters}
    pertalite_reference = {key: value for key, value in reference_data.items() if key in pertalite_meters}

    # Perbandingan untuk solar dan pertalite
    def compare_data(truck_data, reference_data):
        """Bandingkan data dari Excel dengan referensi."""
        comparison_result = {}
        for meter, trucks in reference_data.items():
            excel_trucks = set(truck_data.get(meter, []))  # Data dari Excel (set untuk efisiensi pencarian)
            reference_trucks = set(trucks)  # Data referensi

            # Mencari data yang ada di referensi tapi tidak ada di Excel
            missing_trucks = reference_trucks - excel_trucks
            # Mencari data yang ada di Excel tapi tidak ada di referensi
            additional_trucks = excel_trucks - reference_trucks

            comparison_result[meter] = {
                "missing": list(missing_trucks),
                "additional": list(additional_trucks)
            }
        
        return comparison_result

    solar_comparison = compare_data(solar_meters, solar_reference)
    pertalite_comparison = compare_data(pertalite_meters, pertalite_reference)

    # Menemukan truk yang hilang di satu meter tetapi ada di meter lain
    def find_missing_elsewhere(truck_data, comparison_result):
        """Cari truk yang hilang di satu meter tetapi ditemukan di meter lain."""
        missing_found_elsewhere = []

        for meter, result in comparison_result.items():
            for missing_truck in result["missing"]:
                # Cek apakah nomor truk yang hilang ditemukan di meter lain
                found_in_other_meter = [
                    other_meter for other_meter, data in truck_data.items() if missing_truck in data and other_meter != meter
                ]
                
                if found_in_other_meter:
                    missing_found_elsewhere.append(f"Nomor {missing_truck} mengisi di {meter} sedangkan antrian di {', '.join(found_in_other_meter)}")

        return missing_found_elsewhere

    missing_found_solar = find_missing_elsewhere(solar_meters, solar_comparison)
    missing_found_pertalite = find_missing_elsewhere(pertalite_meters, pertalite_comparison)

    return solar_comparison, pertalite_comparison, missing_found_solar, missing_found_pertalite


@app.route('/')
def index():
    solar_comparison, pertalite_comparison, missing_found_solar, missing_found_pertalite = process_excel()

    # Gabungkan solar dan pertalite dalam satu dictionary
    comparison_result = {**solar_comparison, **pertalite_comparison}
    missing_found_elsewhere = missing_found_solar + missing_found_pertalite  # Gabungkan list

    return render_template(
        'index.html',
        comparison_result=comparison_result, 
        missing_found_elsewhere=missing_found_elsewhere
    )


if __name__ == '__main__':
    app.run(debug=True)
